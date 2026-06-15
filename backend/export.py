import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def export_class_excel(students_data: List[Dict[str, Any]], class_name: str, semester: str) -> bytes:
    """
    Generates an Excel spreadsheet containing class results.
    students_data list items should contain:
      - roll_number, name, subjects (dict of subject_name -> score), total, percentage, grade, rank
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name} Results"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=11)
    
    blue_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark blue
    light_blue_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid") # Medium blue
    zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    # Title Block
    ws.merge_cells("A1:I2")
    title_cell = ws["A1"]
    title_cell.value = f"Class Results: {class_name} ({semester})"
    title_cell.font = title_font
    title_cell.fill = blue_fill
    title_cell.alignment = align_center
    
    # Write Row 3 spacer
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 15
    
    # Identify unique subjects from dataset
    subject_names = []
    for s in students_data:
        for subj in s.get("subjects", {}).keys():
            if subj not in subject_names:
                subject_names.append(subj)
    
    # Headers starting on Row 4
    headers = ["Roll Number", "Name"] + subject_names + ["Total", "Percentage", "Grade", "Rank"]
    ws.row_dimensions[4].height = 25
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Write Data
    row_idx = 5
    for s in students_data:
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        
        # Base student info
        c_roll = ws.cell(row=row_idx, column=1, value=s.get("roll_number"))
        c_name = ws.cell(row=row_idx, column=2, value=s.get("name"))
        c_roll.alignment = align_center
        c_name.alignment = align_left
        
        col_offset = 3
        # Write subjects
        for subj in subject_names:
            score = s.get("subjects", {}).get(subj, "-")
            c_score = ws.cell(row=row_idx, column=col_offset, value=score)
            c_score.alignment = align_center
            col_offset += 1
            
        # Write summaries
        c_total = ws.cell(row=row_idx, column=col_offset, value=s.get("total"))
        c_pct = ws.cell(row=row_idx, column=col_offset+1, value=f"{s.get('percentage'):.1f}%" if isinstance(s.get('percentage'), (int, float)) else s.get('percentage'))
        c_grade = ws.cell(row=row_idx, column=col_offset+2, value=s.get("grade"))
        c_rank = ws.cell(row=row_idx, column=col_offset+3, value=s.get("rank"))
        
        c_total.alignment = align_center
        c_pct.alignment = align_center
        c_grade.alignment = align_center
        c_rank.alignment = align_center
        
        # Apply borders, fonts, Zebra colors
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = normal_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
                
        row_idx += 1
        
    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Save to buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.getvalue()


def export_student_pdf(
    student_info: Dict[str, Any], 
    marks_list: List[Dict[str, Any]], 
    rank: int, 
    percentage: float, 
    grade: str, 
    comments: str
) -> bytes:
    """
    Generates a high-quality PDF Report Card for a single student.
    student_info keys: roll_number, first_name, last_name, class_name, email, contact_number, date_of_birth
    marks_list items: subject_name, marks_obtained, max_marks, grade
    """
    buffer = io.BytesIO()
    
    # 0.5 inch margins
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter, 
        rightMargin=36, 
        leftMargin=36, 
        topMargin=36, 
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Palette styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1A365D'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#2B6CB0'),
        spaceBefore=10,
        spaceAfter=5
    )
    
    cell_style = ParagraphStyle(
        'GridCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#2D3748')
    )
    
    cell_bold_style = ParagraphStyle(
        'GridCellBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    
    label_style = ParagraphStyle(
        'Label',
        parent=cell_style,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#4A5568')
    )

    story = []

    # Title Banner
    story.append(Paragraph("ACADEMIC PROGRESS REPORT CARD", title_style))
    story.append(Spacer(1, 10))

    # Student Details Grid
    student_name = f"{student_info.get('first_name')} {student_info.get('last_name')}"
    details_data = [
        [
            Paragraph("Student Name:", label_style), Paragraph(student_name, cell_style),
            Paragraph("Roll Number:", label_style), Paragraph(student_info.get("roll_number"), cell_style)
        ],
        [
            Paragraph("Class / Section:", label_style), Paragraph(student_info.get("class_name"), cell_style),
            Paragraph("Date of Birth:", label_style), Paragraph(student_info.get("date_of_birth") or "N/A", cell_style)
        ],
        [
            Paragraph("Email Address:", label_style), Paragraph(student_info.get("email") or "N/A", cell_style),
            Paragraph("Contact No:", label_style), Paragraph(student_info.get("contact_number") or "N/A", cell_style)
        ]
    ]
    
    # 7.5 inches printable width (letter is 8.5 x 11, 0.5 in margins means 7.5 in width)
    # Total width in points = 7.5 * 72 = 540 pt
    details_table = Table(details_data, colWidths=[110, 160, 100, 170])
    details_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    
    story.append(details_table)
    story.append(Spacer(1, 15))

    # Academic Marks Table Header
    story.append(Paragraph("Subject-Wise Academic Grades", section_title_style))
    story.append(Spacer(1, 5))
    
    marks_headers = [
        Paragraph("Subject Name", cell_bold_style),
        Paragraph("Marks Obtained", cell_bold_style),
        Paragraph("Max Marks", cell_bold_style),
        Paragraph("Percentage", cell_bold_style),
        Paragraph("Grade", cell_bold_style),
        Paragraph("Status", cell_bold_style)
    ]
    
    marks_data = [marks_headers]
    total_obtained = 0.0
    total_max = 0.0
    
    for m in marks_list:
        score = m.get("marks_obtained", 0)
        max_score = m.get("max_marks", 100)
        total_obtained += score
        total_max += max_score
        
        pct = (score / max_score * 100) if max_score > 0 else 0
        grade_str = m.get("grade", "-")
        status_str = "Pass" if grade_str != "F" else "Fail"
        
        marks_data.append([
            Paragraph(m.get("subject_name"), cell_style),
            Paragraph(str(score), cell_style),
            Paragraph(str(max_score), cell_style),
            Paragraph(f"{pct:.1f}%", cell_style),
            Paragraph(grade_str, cell_bold_style),
            Paragraph(status_str, cell_bold_style if status_str == "Pass" else ParagraphStyle('FailTxt', parent=cell_bold_style, textColor=colors.HexColor('#E53E3E')))
        ])
        
    marks_table = Table(marks_data, colWidths=[180, 80, 80, 80, 60, 60])
    marks_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    
    story.append(marks_table)
    story.append(Spacer(1, 15))
    
    # Results Summary
    story.append(Paragraph("Performance Summary", section_title_style))
    story.append(Spacer(1, 5))
    
    summary_data = [
        [
            Paragraph("Total Marks:", label_style), Paragraph(f"{total_obtained} / {total_max}", cell_style),
            Paragraph("Overall Percentage:", label_style), Paragraph(f"{percentage:.2f}%", cell_style)
        ],
        [
            Paragraph("Final Grade:", label_style), Paragraph(grade, cell_bold_style),
            Paragraph("Class Rank:", label_style), Paragraph(f"Rank {rank}", cell_bold_style)
        ]
    ]
    
    summary_table = Table(summary_data, colWidths=[110, 160, 130, 140])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F7FAFC')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 15))

    # Comments Section
    story.append(Paragraph("Teacher's Remarks & AI Recommendations", section_title_style))
    story.append(Spacer(1, 5))
    
    comment_text = comments or "No comments available for this academic cycle."
    comment_box_style = ParagraphStyle(
        'CommentBox',
        parent=cell_style,
        fontSize=10,
        leading=14
    )
    
    comment_table = Table([[Paragraph(comment_text.replace('\n', '<br/>'), comment_box_style)]], colWidths=[540])
    comment_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,0), colors.HexColor('#F7FAFC')),
        ('GRID', (0,0), (0,0), 0.5, colors.HexColor('#CBD5E0')),
        ('TOPPADDING', (0,0), (0,0), 8),
        ('BOTTOMPADDING', (0,0), (0,0), 8),
        ('LEFTPADDING', (0,0), (0,0), 8),
        ('RIGHTPADDING', (0,0), (0,0), 8),
    ]))
    story.append(comment_table)
    story.append(Spacer(1, 30))
    
    # Signature Lines
    sig_data = [
        [
            Paragraph("________________________<br/>Class Teacher", label_style),
            Paragraph("", label_style),
            Paragraph("________________________<br/>Principal", label_style)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[200, 140, 200])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
    ]))
    story.append(sig_table)

    # Build the document
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes
