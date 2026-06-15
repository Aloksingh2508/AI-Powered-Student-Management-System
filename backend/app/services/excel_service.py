import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_class_excel(students_data: List[Dict[str, Any]], class_name: str, semester: str) -> bytes:
    """
    Generates an Excel spreadsheet containing class results.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name} Results"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=11)
    
    blue_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark blue
    light_blue_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid") # Medium blue
    zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    # Title Block
    ws.merge_cells("A1:I2")
    title_cell = ws["A1"]
    title_cell.value = f"Class Results: {class_name} ({semester})"
    title_cell.font = title_font
    title_cell.fill = blue_fill
    title_cell.alignment = align_center
    
    # Write Row 3 spacer
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 15
    
    # Identify unique subjects from dataset
    subject_names = []
    for s in students_data:
        for subj in s.get("subjects", {}).keys():
            if subj not in subject_names:
                subject_names.append(subj)
    
    # Headers starting on Row 4
    headers = ["Roll Number", "Name"] + subject_names + ["Total", "Percentage", "Grade", "Rank"]
    ws.row_dimensions[4].height = 25
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Write Data
    row_idx = 5
    for s in students_data:
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        
        # Base student info
        c_roll = ws.cell(row=row_idx, column=1, value=s.get("roll_number"))
        c_name = ws.cell(row=row_idx, column=2, value=s.get("name"))
        c_roll.alignment = align_center
        c_name.alignment = align_left
        
        col_offset = 3
        # Write subjects
        for subj in subject_names:
            score = s.get("subjects", {}).get(subj, "-")
            c_score = ws.cell(row=row_idx, column=col_offset, value=score)
            c_score.alignment = align_center
            col_offset += 1
            
        # Write summaries
        c_total = ws.cell(row=row_idx, column=col_offset, value=s.get("total"))
        c_pct = ws.cell(row=row_idx, column=col_offset+1, value=f"{s.get('percentage'):.1f}%" if isinstance(s.get('percentage'), (int, float)) else s.get('percentage'))
        c_grade = ws.cell(row=row_idx, column=col_offset+2, value=s.get("grade"))
        c_rank = ws.cell(row=row_idx, column=col_offset+3, value=s.get("rank"))
        
        c_total.alignment = align_center
        c_pct.alignment = align_center
        c_grade.alignment = align_center
        c_rank.alignment = align_center
        
        # Apply borders, fonts, Zebra colors
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = normal_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
                
        row_idx += 1
        
    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Save to buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.getvalue()
